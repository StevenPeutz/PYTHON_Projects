#connecting to newsAPI
#while using textblob for NLP sentiment analysis

#reqs:
# pip install newsapi-python
# pip install -U textblob
from newsapi import NewsApiClient
import seaborn as sns
import matplotlib.pyplot as plt
import csv
from nltk.util import Index
import pandas as pd
from textblob import TextBlob
from dateutil import parser
api = NewsApiClient(api_key='d7be4b06f0a94595825b3137e3683ad4')

## documentenation:  https://newsapi.org/docs/client-libraries/python
btc_content = api.get_everything(q='bitcoin',
                             sources = 'bbc-news, metro, financial-times, cnbc, business-insider, reuters, wired, vice-news, usa-today, time, the-washington-times, the-washington-post, the-wallstreet-journaal, the-verge, the-next-web, the-huffington-post, the-hill, the-american-conservative, techradar, techcrunch, reddit-r-all, recode, polygon, politico, nfl-news, new-york-magazine, new-scientist',
                             language = 'en'
                             #from_param='2020-12-01',
                             #         to='2021-02-12'
)
#print(content)
btc_content_df = pd.DataFrame(btc_content['articles'])


btc_content_df['date'] = btc_content_df.apply(lambda x: parser.parse(x['publishedAt']).strftime('%Y.%m.%d'), axis=1)
#content_df['time'] = content_df.apply(lambda x: parser.parse(x['publishedAt']).strftime('%H:%M'), axis=1)
btc_content_df['polarity'] = btc_content_df.apply(lambda x: TextBlob(x['description']).sentiment.polarity, axis=1)
#content_df['subjectivity'] = content_df.apply(lambda x: TextBlob(x['description']).sentiment.subjectivity, axis=1)


#########################################
##### testing some words ##############
#from textblob.sentiments import NaiveBayesAnalyzer
#alt = TextBlob('this thing is great', analyzer=NaiveBayesAnalyzer())
alt = TextBlob('this thing is great')
alt.sentiment
#learnings: the pattern analyzer misses a lot like 'to the moon' and 'will rise', and 'rally', bullish and bearish etc 





#appending
#creating new file if missing
#to be solved are;
#   - only unique rows should be added
#   - header should be added upon creation of xls and only upon creation.


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

append_df_to_excel(r'XLS_files/test444.xlsx', btc_content_df, header=None, index=False)










###############
# next steps would be to find a good weight (e.g. number of readers) to create a weigthed average of positivty scores
# and then plot it against bitcoin historical price data
btc_content_df.head(1)
btc_content_df.shape

#content_df.groupby('date')['polarity'].mean().plot(kind='line')
datanew = btc_content_df.groupby(by = 'date').mean()
datanew.plot(kind='line')
plt.show()

datanew.to_excel(r'XLS_files/test_datanew.xlsx', index = True)

